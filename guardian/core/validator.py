"""
Preflight Validation - Ensures project readiness before monitoring starts
"""
import json
from pathlib import Path
from typing import Dict, List, Tuple, Any
import logging

logger = logging.getLogger("DRE_Validator")


class ValidationError:
    """Represents a blocking validation issue"""
    def __init__(self, category: str, title: str, message: str, fix: str):
        self.category = category
        self.title = title
        self.message = message
        self.fix = fix
        self.blocking = True


class ValidationWarning:
    """Represents a non-blocking validation warning"""
    def __init__(self, category: str, title: str, message: str, recommendation: str):
        self.category = category
        self.title = title
        self.message = message
        self.recommendation = recommendation
        self.blocking = False


class PreflightValidator:
    """
    Authoritative preflight validation for DRE projects
    Used by both 'doctor' and 'monitor' to ensure consistency
    """
    
    @staticmethod
    def validate_project(manifest_path: str) -> Tuple[List[ValidationError], List[ValidationWarning]]:
        """
        Comprehensive preflight check
        
        Returns:
            (errors, warnings) - errors block monitoring, warnings are informational
        """
        errors = []
        warnings = []
        
        manifest_path_obj = Path(manifest_path)
        
        # Check 1: Manifest file exists
        if not manifest_path_obj.exists():
            errors.append(ValidationError(
                category="Manifest",
                title="Manifest file not found",
                message=f"The manifest file does not exist:\n  {manifest_path}",
                fix="Run 'init' to create project scaffolding, then configure manifest.json"
            ))
            return errors, warnings  # Can't continue without manifest
        
        # Check 2: Manifest JSON is valid
        try:
            with open(manifest_path_obj, 'r', encoding='utf-8') as f:
                manifest_data = json.load(f)
        except json.JSONDecodeError as e:
            errors.append(ValidationError(
                category="Manifest",
                title="Manifest contains invalid JSON",
                message=f"The manifest file has syntax errors:\n  Line {e.lineno}: {e.msg}",
                fix="Open manifest.json in a text editor and fix the JSON syntax error"
            ))
            return errors, warnings  # Can't continue with invalid JSON
        
        # Check 3: Required manifest fields
        required_fields = ["project_id", "project_name", "target_file", "assertions"]
        missing_fields = [f for f in required_fields if f not in manifest_data]
        if missing_fields:
            errors.append(ValidationError(
                category="Manifest",
                title="Manifest missing required fields",
                message=f"These required fields are missing: {', '.join(missing_fields)}",
                fix="Add the missing fields to your manifest.json"
            ))
            return errors, warnings
        
        # Check 4: Excel file exists
        target_file = manifest_data["target_file"]
        
        # Resolve relative paths
        if not Path(target_file).is_absolute():
            manifest_dir = manifest_path_obj.parent
            excel_path = manifest_dir / target_file
        else:
            excel_path = Path(target_file)
        
        if not excel_path.exists():
            errors.append(ValidationError(
                category="Excel File",
                title="Excel file not found",
                message=f"The manifest refers to an Excel file that does not exist:\n  {excel_path}\n\nYour manifest specifies: target_file = \"{target_file}\"",
                fix=f"Place your Excel file in the correct location, or update\nthe 'target_file' field in manifest.json to match your file's actual name"
            ))
            return errors, warnings  # Can't validate sheets/cells without Excel file
        
        # Check 5: Excel file is readable
        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
        except PermissionError:
            errors.append(ValidationError(
                category="Excel File",
                title="Excel file is locked",
                message=f"The Excel file is currently open in another application:\n  {excel_path.name}",
                fix="Close the Excel file and try again"
            ))
            return errors, warnings
        except Exception as e:
            errors.append(ValidationError(
                category="Excel File",
                title="Excel file cannot be opened",
                message=f"Failed to open Excel file:\n  {excel_path.name}\n  Error: {str(e)}",
                fix="Ensure the file is a valid Excel workbook (.xlsx format)"
            ))
            return errors, warnings
        
        # Check 6: Assertions are defined
        assertions = manifest_data.get("assertions", [])
        if not assertions:
            warnings.append(ValidationWarning(
                category="Manifest",
                title="No assertions defined",
                message="The manifest has no assertions to monitor",
                recommendation="Add assertions to manifest.json to define what cells to govern"
            ))
            return errors, warnings  # No point checking cells if no assertions
        
        # Check 7: Validate each assertion's binding (sheet and cell exist)
        for idx, assertion in enumerate(assertions):
            assertion_id = assertion.get("id", f"assertion_{idx}")
            
            # Check binding structure
            if "binding" not in assertion:
                errors.append(ValidationError(
                    category="Assertion",
                    title=f"Assertion '{assertion_id}' missing binding",
                    message=f"Assertion '{assertion_id}' does not specify which cell to monitor",
                    fix=f"Add a 'binding' field with 'sheet' and 'cell' for assertion '{assertion_id}'"
                ))
                continue
            
            binding = assertion["binding"]
            sheet_name = binding.get("sheet")
            cell_ref = binding.get("cell")
            
            if not sheet_name:
                errors.append(ValidationError(
                    category="Assertion",
                    title=f"Assertion '{assertion_id}' missing sheet name",
                    message=f"The binding for '{assertion_id}' does not specify a sheet",
                    fix=f"Add 'sheet' field to the binding for assertion '{assertion_id}'"
                ))
                continue
            
            if not cell_ref:
                errors.append(ValidationError(
                    category="Assertion",
                    title=f"Assertion '{assertion_id}' missing cell reference",
                    message=f"The binding for '{assertion_id}' does not specify a cell",
                    fix=f"Add 'cell' field to the binding for assertion '{assertion_id}'"
                ))
                continue
            
            # Check sheet exists
            if sheet_name not in wb.sheetnames:
                available_sheets = "', '".join(wb.sheetnames[:5])  # Show first 5 sheets
                errors.append(ValidationError(
                    category="Excel Sheet",
                    title=f"Sheet '{sheet_name}' not found",
                    message=f"Assertion '{assertion_id}' refers to sheet '{sheet_name}',\nbut that sheet does not exist in your Excel file.\n\nAvailable sheets: '{available_sheets}'{' ...' if len(wb.sheetnames) > 5 else ''}",
                    fix=f"Either:\n  • Open your Excel file and check the exact sheet name (case-sensitive)\n  • Update the 'sheet' field in assertion '{assertion_id}' to match"
                ))
                continue
            
            # Check cell is readable
            try:
                ws = wb[sheet_name]
                cell_value = ws[cell_ref].value
                
                # Warn if cell is empty
                if cell_value is None:
                    warnings.append(ValidationWarning(
                        category="Cell Value",
                        title=f"Cell '{sheet_name}!{cell_ref}' is empty",
                        message=f"Assertion '{assertion_id}' monitors an empty cell",
                        recommendation="Ensure this is intentional, or update the cell reference"
                    ))
                
                # Warn if expecting numeric but got text
                if "baseline_value" in assertion and isinstance(assertion["baseline_value"], (int, float)):
                    if cell_value is not None and not isinstance(cell_value, (int, float)):
                        warnings.append(ValidationWarning(
                            category="Cell Type",
                            title=f"Cell '{sheet_name}!{cell_ref}' contains non-numeric value",
                            message=f"Assertion '{assertion_id}' expects a number, but the cell contains: {type(cell_value).__name__}",
                            recommendation="Verify the cell reference is correct"
                        ))
                        
            except Exception as e:
                errors.append(ValidationError(
                    category="Cell Reference",
                    title=f"Invalid cell reference '{cell_ref}'",
                    message=f"Assertion '{assertion_id}' has an invalid cell reference:\n  Sheet: {sheet_name}\n  Cell: {cell_ref}\n  Error: {str(e)}",
                    fix=f"Use a valid Excel cell reference (e.g., 'A1', 'B12', 'AA5')"
                ))
                continue
        
        wb.close()
        return errors, warnings
