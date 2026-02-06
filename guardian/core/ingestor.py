import time
import openpyxl
import hashlib
import logging
from typing import Dict, Any
from pathlib import Path
from .schema import DREManifest

logger = logging.getLogger("DREIngestor")

class ExcelIngestor:
    def __init__(self, manifest: DREManifest, manifest_path: str = None):
        self.manifest = manifest
        # If manifest_path is provided, resolve target_file relative to manifest directory
        if manifest_path:
            manifest_dir = Path(manifest_path).parent
            self.target_path = str(manifest_dir / manifest.target_file)
        else:
            self.target_path = manifest.target_file

    def _get_formula_hash(self, formula: str) -> str:
        """Generates a fingerprint of the cell logic to detect hijacking."""
        if not formula or not str(formula).startswith('='):
            return "static_value"
        return hashlib.sha256(str(formula).encode()).hexdigest()

    def read_data(self, retries=5, delay=0.5) -> Dict[str, Any]:
        """
        Reads the target file with backoff logic to handle file locks.
        Fulfills FR5 (Resilience) and Edge Case A.
        """
        for i in range(retries):
            try:
                # PASS A: Get Values
                wb_data = openpyxl.load_workbook(
                    self.target_path, data_only=True, read_only=True
                )
                # PASS B: Get Formulas (for hijack detection)
                wb_logic = openpyxl.load_workbook(
                    self.target_path, data_only=False, read_only=True
                )
                
                results = {}
                sheet_data = wb_data.active
                sheet_logic = wb_logic.active

                for assertion in self.manifest.assertions:
                    cell_coord = assertion.binding.cell
                    sheet_name = assertion.binding.sheet
                    
                    # Get the correct sheet (default to active if not specified)
                    if sheet_name:
                        try:
                            if sheet_name not in wb_data.sheetnames:
                                raise ValueError(
                                    f"Sheet '{sheet_name}' not found in Excel file.\n"
                                    f"Assertion '{assertion.id}' cannot be evaluated.\n"
                                    f"Available sheets: {', '.join(wb_data.sheetnames[:5])}"
                                )
                            
                            sheet_data = wb_data[sheet_name]
                            sheet_logic = wb_logic[sheet_name]
                        except Exception as e:
                            logger.error(f"Sheet access error for assertion '{assertion.id}': {e}")
                            raise
                    # else: use active sheet already set above
                    
                    # Get cell value and formula
                    try:
                        val = sheet_data[cell_coord].value
                        formula = sheet_logic[cell_coord].value
                        current_hash = self._get_formula_hash(formula)

                        results[assertion.id] = {
                            "value": val,
                            "formula_hash": current_hash,
                            "logical_name": assertion.logical_name
                        }
                    except Exception as e:
                        logger.error(f"Cell access error for assertion '{assertion.id}': {e}")
                        raise ValueError(
                            f"Cannot read cell '{sheet_name}!{cell_coord}' for assertion '{assertion.id}'.\n"
                            f"Verify the cell reference is valid."
                        )
                
                wb_data.close()
                wb_logic.close()
                return results

            except PermissionError:
                logger.warning(f"File locked by Excel. Retry {i+1}/{retries}...")
                time.sleep(delay * (2 ** i)) # Exponential backoff
            except ValueError:
                # Re-raise validation errors immediately (don't retry)
                raise
            except FileNotFoundError:
                raise FileNotFoundError(
                    f"Excel file not found: {self.target_path}\n"
                    f"Verify the 'target_file' path in manifest.json is correct."
                )
            except Exception as e:
                logger.error(f"Ingestion error: {e}", exc_info=True)
                raise RuntimeError(
                    f"Failed to read Excel file:\n{str(e)}\n\n"
                    f"Ensure the file is a valid Excel workbook (.xlsx format)."
                )

        raise PermissionError(
            f"Excel file is locked after {retries} attempts.\n"
            f"Close the file in Excel and try again."
        )